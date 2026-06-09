import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


CATEGORY_NAMES = {
    1: "Ăn uống",
    2: "Di chuyển",
    3: "Mua sắm",
    5: "Giải trí",
    6: "Sức khỏe",
    7: "Khác",
}


def load_labels(path: Path) -> list[int]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return [int(row["category_id"]) for row in csv.DictReader(file, delimiter=";")]


def update_workbook(source: Path, destination: Path, labels: list[int]) -> None:
    workbook = load_workbook(source)
    sheet = workbook["Review_Data"]
    headers = [cell.value for cell in sheet[1]]
    category_id_column = headers.index("category_id") + 1
    category_name_column = headers.index("category_name") + 1

    if sheet.max_row - 1 != len(labels):
        raise ValueError(f"{source} has {sheet.max_row - 1} rows, expected {len(labels)}")

    for row_number, category_id in enumerate(labels, start=2):
        sheet.cell(row_number, category_id_column).value = category_id
        sheet.cell(row_number, category_name_column).value = CATEGORY_NAMES[category_id]

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", required=True, type=Path)
    parser.add_argument("--single-source", required=True, type=Path)
    parser.add_argument("--mix-source", required=True, type=Path)
    parser.add_argument("--single-output", required=True, type=Path)
    parser.add_argument("--mix-output", required=True, type=Path)
    parser.add_argument("--single-rows", default=900, type=int)
    args = parser.parse_args()

    labels = load_labels(args.csv)
    single_labels = labels[: args.single_rows]
    mix_labels = labels[args.single_rows :]
    update_workbook(args.single_source, args.single_output, single_labels)
    update_workbook(args.mix_source, args.mix_output, mix_labels)
    print(f"single_rows={len(single_labels)}, mix_rows={len(mix_labels)}")


if __name__ == "__main__":
    main()
